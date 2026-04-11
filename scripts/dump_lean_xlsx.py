"""
Tiện ích dev tùy chọn: đọc một workbook .xlsx và dump nội dung ra text.
Ứng dụng PPMS không phụ thuộc Excel; script chỉ phục vụ khảo sát/thiết kế thủ công nếu cần.
"""
import openpyxl
import sys

path = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\ASUS\Desktop\lean_eval_v2.xlsx"
out_path = sys.argv[2] if len(sys.argv) > 2 else "storage/app/lean_dump.txt"

wb = openpyxl.load_workbook(path, data_only=True)
lines = []
for name in wb.sheetnames:
    ws = wb[name]
    lines.append(f"=== {name} (rows {ws.max_row}, cols {ws.max_column}) ===")
    mr = min(ws.max_row or 1, 100)
    mc = min(ws.max_column or 1, 35)
    for ri in range(1, mr + 1):
        row = [ws.cell(ri, cj).value for cj in range(1, mc + 1)]
        lines.append(f"{ri}\t{row}")
    lines.append("")

with open(out_path, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("Wrote", out_path)
